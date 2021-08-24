#--------------------IMPORTAR LIBRERIA------------------
from openpyxl import load_workbook
import csv
################################----CREAR Y ESCRIBIR DOCUMENTO-------##################################
def crea_cargar(arch_fuente,arch_creado):
    #---------------------------------Cargar archivo-------------------------------------
    editor = load_workbook(arch_fuente)
    conteo = editor.active #CREARO OBJETO PARA MANIPULACIÓN DE DATOS
    num_fil = conteo.max_row#obtengo numero usado de filas
    num_col = conteo.max_column#obtengo numer usado de columnas
    tabla_datos = [[conteo.cell(row=fil,column=col).value for col in range(1,num_col-2)]  for fil in range(5,num_fil+1)]
    #-------------Crear archivo CSV y escribir---------------------
    escribir_archivo(arch_creado,tabla_datos)
####################################---LECTURA DEL DOCUMENTO---#####################################
def extraer_lista_archivo(arch_leer):
    with open(arch_leer) as f:
        reader = csv.reader(f,delimiter=';')
        doc_leer = [[row[0],row[1],row[2],row[3],row[4],row[5],row[6]] for row in reader ]
    f.close()
    return doc_leer
####################################---ESCRIBIR EN DOCUMENTO---#####################################
def escribir_archivo(archivo_escribir,lista_datos):
    archivo = open(archivo_escribir,"w",newline='')
    escribir = csv.writer(archivo,delimiter=';')
    escribir.writerows(lista_datos)
    archivo.close()
#####################################-----LECTURA Y MODIFICACIÓN DE CSV------########################
def modificar(archivo,fila,columna,valor):
    lista = extraer_lista_archivo(archivo)
    lista[fila][columna] = valor 
    escribir_archivo(archivo,lista)
####################################------INICAR PRUEBA-----######################################
documento = 'Existencia_al_260620_(2).xlsx'
arch = 'Resultado.csv'
crea_cargar(documento,arch)
for a in range(len(extraer_lista_archivo(arch))):
    modificar(arch,a,0,a)